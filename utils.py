from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from utils import *
import pandas as pd
import re

def add_round(num_rounds, sheet):
    for col in range(1, num_rounds + 1):
        column_letter = get_column_letter(col)  # Get the corresponding column letter
        cell_address = f"{column_letter}5"  # Row 5 for all columns
        sheet[cell_address] = f"Round {col}"

def sort_data_table_by_column(data_table, column_index):
    sorted_data_table = sorted(data_table, key=lambda row: row[column_index], reverse=True)
    return sorted_data_table


