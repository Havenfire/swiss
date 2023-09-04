from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from utils import *
import pandas as pd


filename = "C://Users//bsliu//Downloads//SIGN UP_ Emory March 2023 Chess Tournament (Responses).xlsx"

#load excel file
workbook = load_workbook(filename)
#workbook.create_sheet('Swiss Tournament')

#open workbook
sheet_form = workbook['Form Responses 1']
sheet_tournament = workbook['Swiss Tournament']

data_table = []

# Iterate through rows (excluding the header row)
for row in sheet_form.iter_rows(min_row=2, values_only=True):
    row_data = {}
    for col_idx, cell_value in enumerate(row, start=1):
        header_cell = sheet_form.cell(row=1, column=col_idx)
        header_text = header_cell.value
        
        if header_text in ['What is your full name?', 'Rating?']:  # Include only specific columns
            row_data[header_text] = cell_value
    data_table.append(row_data)


# sort_data_table_by_column(sheet_form, 5)

for row in data_table:
    print(row)


sheet_tournament["A1"] = "Welcome to Emory's Chess Swiss Tournament Program"
sheet_tournament["A2"] = "Built by Blake Liu"

num_rounds = 1
add_round(num_rounds = num_rounds, sheet = sheet_tournament)





#save the file
workbook.save(filename)